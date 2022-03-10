from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.core.validators import FileExtensionValidator
from django.shortcuts import render
from openpyxl import load_workbook

from account import utils
from account.forms import ProfileEditForm, UserEditForm, UserRegistrationForm

from .models import Profile


def register(request):
    """View for registration."""
    if request.method == 'POST':
        user_form = UserRegistrationForm(request.POST)
        if user_form.is_valid():
            if 'registeruser' in request.POST:
                new_user = user_form.save(commit=False)
                generated_password = utils.get_random_password()
                new_user.set_password(generated_password)
                new_user.save()
                Profile.objects.create(user=new_user)
                cd = user_form.cleaned_data
                email_list = []
                email_list.append(cd['email'])
                subject = 'SWTPZ konto użytkownika {} {}'.format(cd['first_name'], cd['last_name'])
                message = 'Konto użytkownika {} {} zostało założone w SWTPZ\nOto dane potrzebne do logowania: \n' \
                          'Login: {}\nHasło: {}'.format(cd['first_name'], cd['last_name'], new_user, generated_password)
                send_mail(subject, message, 'zswtpz@gmail.com', email_list)
                return render(request, 'dashboard/register_complete.html',
                              {})
            elif 'uploadfile' in request.POST and request.FILES['filewithdetails']:
                file = request.FILES['filewithdetails']
                valid = FileExtensionValidator(['xls', 'xlsx'])
                if valid(file) != None:
                    error = "Nieprawidłowy format pliku!"
                    return render(request, 'dashboard/registeruser.html',
                                  {'user_form': user_form, 'error': error})
                wb = load_workbook(filename=BytesIO(file.read()))
                ws = wb.active
                account_list = []

                for row in ws.iter_rows():
                    account_list.append([row[0].value, row[1].value, row[2].value])

                for account in account_list:
                    new_user = User(username=utils.create_username(account[0], account[1]), email=account[2],
                                    first_name=account[0], last_name=account[1])
                    new_user.save()
                    generated_password = utils.get_random_password()
                    new_user.set_password(generated_password)
                    Profile.objects.create(user=new_user)
                    email_list = []
                    email_list.append(account[2])
                    subject = 'SWTPZ konto użytkownika {} {}'.format(account[0], account[1])
                    message = 'Konto użytkownika {} {} zostało założone w SWTPZ\nOto dane potrzebne do logowania: \n' \
                              'Login: {}\nHasło: {}'.format(account[0], account[1], new_user,
                                                            generated_password)
                    send_mail(subject, message, 'zswtpz@gmail.com', email_list)
                succes = "Konta zostały pomyślnie założone!"
                return render(request, 'dashboard/registeruser.html',
                              {'user_form': user_form, 'succes': succes})
    else:
        user_form = UserRegistrationForm()
    return render(request, 'dashboard/registeruser.html',
                  {'user_form': user_form})


@login_required
def edit_profile(request):
    """View for editting user profile."""
    if request.method == "POST":
        user_form = UserEditForm(instance=request.user, data=request.POST)
        profile_form = ProfileEditForm(
            instance=request.user.profile,
            data=request.POST
        )
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
    else:
        user_form = UserEditForm(instance=request.user)
        profile_form = ProfileEditForm(instance=request.user.profile)
    return render(request,
                  'account/account_details.html',
                  {'user_form': user_form,
                   'profile_form': profile_form})
